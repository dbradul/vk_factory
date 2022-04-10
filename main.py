import os
import random

import sys
import time
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from models import VkClientProxy
from utils import logger

load_dotenv()

NUM_USERS_BEFORE_ASK = int(os.getenv('NUM_USERS_BEFORE_ASK'))
MIN_WAIT = int(os.getenv('MIN_WAIT'))
MAX_WAIT = int(os.getenv('MAX_WAIT'))


def repost(client, post):
    wall_post_id = post.split('/')[-1]
    resp = client._obj._vk.method('wall.repost', {'object': wall_post_id})
    logger.info(f'Reposted wall post with id={wall_post_id} with result {resp}')


def add_friend(client, id, invite_msg):
    resp = client._obj._vk.method('friends.add', {'user_id': int(id), 'text': invite_msg})
    logger.info(f'Added friend with id={id} with result {resp}')


def process_sheet(client, sheet, invite_msg, num_friends):
    row_count = min(num_friends, sheet.max_row)
    success_rows = 0

    headers = {
        f'{letter}': sheet[f'{letter}1'].value
        for letter in 'ABCDEF'
    }

    for idx, r in enumerate(range(2, row_count + 1), 1):
        try:
            value = sheet[f'D{str(r)}'].value
            if headers['D'].lower() == 'id':
                add_friend(client, int(value), invite_msg)
            elif headers['D'].lower() == 'urltorepost':
                repost(client, value)

            success_rows += 1

            wait_slot = random.randint(MIN_WAIT, MAX_WAIT)
            logger.info(f'Sleeping for {wait_slot}s...')
            time.sleep(wait_slot)

        except Exception as ex:
            logger.error(f'Failed to add friend {id}: {ex}')

        finally:
            if (idx % NUM_USERS_BEFORE_ASK) == 0:
                print('\n---------------------------------------------------------')
                print('EXCEEDED MAX USER LIMIT')
                input('PLEASE, PRESS [ENTER] TO CONTINUE...')

    return f'{success_rows}/{row_count}'


def process_file(client, filename):
    wb = load_workbook(filename)
    ws: Worksheet = wb.active
    sheet = wb.worksheets[0]
    sheet_map = {sheet.title: sheet for sheet in wb.worksheets}
    row_count = sheet.max_row
    prev_login = None
    for r in range(2, row_count + 1):
        try:
            login = str(ws[f'B{str(r)}'].value)
            sheetname = ws[f'C{str(r)}'].value
            invite_msg = ws[f'D{str(r)}'].value
            num_friends = ws[f'E{str(r)}'].value

            sheet = sheet_map.get(sheetname)
            if not sheet:
                raise RuntimeError(f'Sheet with name {sheetname} is not found')
            if not prev_login or prev_login != login:
                logins = [account[0] for account in client._accounts]
                if login not in logins:
                    raise RuntimeError(f'Account {login} is not found among those: {logins}')
                else:
                    # client.auth_as(login)
                    client.direct_auth(login, app_id=os.getenv('VK_APP_ID'), client_secret=os.getenv('VK_APP_SECRET'))
                prev_login = login

            logger.info(f'Starting {login} for audience {sheetname}...')
            resp = process_sheet(client, sheet, invite_msg, num_friends)
            logger.info(f'Finished {login} for audience {sheetname}')
            # time.sleep(random.randint(MIN_WAIT, MAX_WAIT))
        except Exception as ex:
            logger.error(f'Failed to send message for account {login}: {ex}')

    wb.save(filename)


def main():
    vk_client = VkClientProxy()
    vk_client.load_accounts()
    # vk_client.auth()

    if len(sys.argv) > 1:
        param = sys.argv[1]
        process_file(vk_client, param)


if __name__ == '__main__':
    main()
